import csv
import sys
import pandas as pd
import numpy as np
import geopandas as gpd
from typing import NamedTuple
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import jellyfish
import math

SF_CENSUS_PATH = (
    Path(__file__).parent.parent / "raw-data/census/sf_census_tracts_2020.csv"
)
CALI_TRACTS_SHP = (
    Path(__file__).parent.parent
    / "raw-data/census/cali_tracts_shapefiles/tl_2025_06_tract.shp"
)

POP_PATH = (
    Path(__file__).parent.parent / "raw-data/census/acs_sf_population_2020_24.csv"
)
RENT_PATH = (
    Path(__file__).parent.parent / "raw-data/census/acs_sf_median_rent_2020_24.csv"
)

REPORT_PATH = (
    Path(__file__).parent.parent / "raw-data"/ "311_cases.csv"
)
ENCAMP_PATH = (
      Path(__file__).parent.parent / "raw-data"/ "encampments_counts.xlsx"
)

HH_INC_PATH = (
    Path(__file__).parent.parent / "raw-data/census/acs_sf_median_hh_income_2020_24.csv"
)
RACE_PATH = Path(__file__).parent.parent / "raw-data/census/acs_sf_race_2020_24.csv"
RENTER_UNITS_PATH = Path(__file__).parent.parent / "raw-data/census/acs_sf_housing_units_2020_24.csv"

SF_ACS_JOIN = Path(__file__).parent.parent / "clean-data/census_acs_join.csv"
SF_TRACTS_SHP = Path(__file__).parent.parent / "clean-data/sf_shapefiles/sf_tracts.shp"
MERGED_SF_TRACTS_SHP = (
    Path(__file__).parent.parent
    / "clean-data/merged_sf_shapefiles/merged_sf_tracts.shp"
)

POP_ID = "AUO6E001"
RENT_ID = "AUWGE001"
HH_INC_ID = "AURUE001"
WHITE_POP_ID = "AUO7E002"
RENTER_UNITS_ID = "AUUEE003"

EXCLUDE_GEOIDS = ["06075980401", "06075980200"]


class Encampment(NamedTuple):
    ### unique  encampmemnt id per quarter
    id: int
    tents: int
    structures: int
    vehicles: int

    year: int
    month: int
    date_time: datetime
    lat: float
    lon: float
    neighborhood: str


class EncampmentReport(NamedTuple):
    id: int
    year: int
    month: int
    date_time: datetime
    lat: float
    lon: float
    neighborhood: str
    


def rate(score):
    if score >= 0.95:
        return "high"
    if score < 0.95 and score >= 0.80:
        return "medium"
    return "low"


## Clean 311 data
def clean_311():

    file_input = REPORT_PATH

    with open(file_input, newline="") as csvfile:
        """
        Given a CSV containing 311, return a list of Encampment report objects.
        """
        output_report = []
        reader = csv.DictReader(csvfile)
        for row in reader:
            ### Clean the date ####
            datetime_str = row.get("Opened").replace(" PM", "")
            datetime_str = datetime_str.replace(" AM", "")
            datetime_object = datetime.strptime(datetime_str, "%m/%d/%Y %H:%M:%S")
            date_year = datetime_object.year
            date_month = datetime_object.month
            if row['Latitude'] == '':
                lat = 0 
            else:
                lat = float(row['Latitude'])
                            
            if row['Longitude'] == '':
                lon = 0
            else:
                lon = float(row['Longitude'])

            tuple_out = EncampmentReport(
                row.get("CaseID"),
                date_year,
                date_month,
                datetime_object,
                lat,
                lon,
                row.get("Neighborhood")
            )
            ### Remove reports with missing geographic information ####
            tolerance = .01
            missing_location = math.isclose(tuple_out.lat, 0.0, abs_tol=tolerance) and math.isclose(tuple_out.lon, 0.0, abs_tol=tolerance)
            if not missing_location:
                output_report.append(tuple_out)

    return output_report


### Clean encampment data ###
def clean_encampment():
    file_input = ENCAMP_PATH
    wb = load_workbook(file_input)
    sheet_obj = wb.active

    for i in range(1, sheet_obj.max_column + 1):
        print(sheet_obj.cell(row=1, column=i).value)

    for i in range(1, sheet_obj.max_column + 1):
        print(sheet_obj.cell(row=2, column=i).value)

    assert sheet_obj.cell(row=2, column=3).value == "Tents"
    assert sheet_obj.cell(row=2, column=4).value == "Structures"
    assert sheet_obj.cell(row=2, column=5).value == "Passenger Vehicles"
    assert sheet_obj.cell(row=2, column=6).value == "Other Vehicles"
    assert sheet_obj.cell(row=2, column=8).value == "Neighborhood"
    assert sheet_obj.cell(row=2, column=10).value == "Latitude"
    assert sheet_obj.cell(row=2, column=11).value == "Longitude"

    output_encampment = []
    for i in range(3, sheet_obj.max_row + 1):
        sheet_obj.cell(row=3, column=1).value
        date_obj = sheet_obj.cell(row=i, column=1).value
        date_string = date_obj.strftime("%m/%d/%Y")
        tents = sheet_obj.cell(row=i, column=3).value
        structure = sheet_obj.cell(row=i, column=4).value
        vehicles = (
            sheet_obj.cell(row=i, column=5).value
            + sheet_obj.cell(row=i, column=6).value
        )
        neighborhood = sheet_obj.cell(row=i, column=8).value

        lat = float(sheet_obj.cell(row=i, column=10).value)
        lon = float(sheet_obj.cell(row=i, column=11).value)
        obj = Encampment(i, tents,
        structure,
        vehicles,
        date_obj.year,
        date_obj.month,
        date_string,
        lat,
        lon,
        neighborhood)
        output_encampment.append(obj)
    return output_encampment


#### Merge the two files to filter out 311 reports associatd with marked/observed encampments ####


def attached_311_reports(output_encampment, output_report):

    associated_encamp = []
    year_2021 = [encamp for encamp in output_encampment if encamp.year == 2021]
    report_2021 = [report for report in output_report if report.year == 2021]
    month_dec_2020 = [report for report in output_report if report.year == 2020 and report.month == 12]
    month_jan_2022 = [report for report in output_report if report.year == 2022 and report.month == 1]
    report_2021.extend(month_dec_2020)
    report_2021.extend(month_jan_2022)

    for encampment in year_2021:

            for report in report_2021:
                format_pattern = '%m/%d/%Y'
                diff =  datetime.strptime(encampment.date_time,format_pattern) - report.date_time
                if abs(diff.days) <= 15:
                    point1 = (encampment.lat, encampment.lon)
                    point2 = (report.lat, report.lon)
                    if (
                        rate(
                            jellyfish.jaro_winkler_similarity(
                                encampment.neighborhood.lower(), report.neighborhood.lower()
                            )
                        )
                        == "high"):
                            if (distance.distance(point1, point2).miles) < 0.2:
                                associated_encamp.append((encampment, report))

### Bounding box 
### neighborhood bound
# ### For each encampment, create a bounding box for each encampmemnt  
### use the timeit library 
### timeit helpful to see which functions are taking the longest

def process_acs_data():
    """
    Load the data from the ACS files saved, impute negative or zero values in
    rent and household income, merge them based on GeoID, and save the merged
    dataframe to a new file.

    In each ACS file:
    - Tracts will be identified by their GeoID ("TL_GEO_ID")
    - Data to be merged will be found in the column with the unique identifier
    """
    csv.field_size_limit(sys.maxsize)

    pop_df = pd.read_csv(
        POP_PATH, usecols=["TL_GEO_ID", POP_ID], dtype={"TL_GEO_ID": "str"}
    )
    race_df = pd.read_csv(
        RACE_PATH, usecols=["TL_GEO_ID", WHITE_POP_ID], dtype={"TL_GEO_ID": "str"}
    )
    rent_df = pd.read_csv(
        RENT_PATH, usecols=["TL_GEO_ID", RENT_ID], dtype={"TL_GEO_ID": "str"}
    )
    hh_inc_df = pd.read_csv(
        HH_INC_PATH, usecols=["TL_GEO_ID", HH_INC_ID], dtype={"TL_GEO_ID": "str"}
    )
    renter_df = pd.read_csv(
        RENTER_UNITS_PATH, usecols=["TL_GEO_ID", RENTER_UNITS_ID], dtype={"TL_GEO_ID": "str"}
    )

    # Impute negative or zero values in rent and household income dataframes
    # with the mean of their positive values
    mean_rent = round(rent_df.loc[rent_df[RENT_ID] > 0, RENT_ID].mean())
    rent_df.loc[rent_df[RENT_ID] <= 0, RENT_ID] = mean_rent

    mean_hh_inc = round(hh_inc_df.loc[hh_inc_df[HH_INC_ID] > 0, HH_INC_ID].mean())
    hh_inc_df.loc[hh_inc_df[HH_INC_ID] <= 0, HH_INC_ID] = mean_hh_inc

    # Rename population, race, rent, and household income column names
    pop_df = pop_df.rename(columns={POP_ID: "population"})
    race_df = race_df.rename(columns={WHITE_POP_ID: "white_pop"})
    rent_df = rent_df.rename(columns={RENT_ID: "med_rent"})
    hh_inc_df = hh_inc_df.rename(columns={HH_INC_ID: "med_hh_inc"})
    renter_df = renter_df.rename(columns={RENTER_UNITS_ID: "rent_units"})

    # Merge individual dataframes based on GEO_ID
    joined_df = pd.merge(pop_df, race_df, on="TL_GEO_ID", how="left")
    joined_df = pd.merge(joined_df, rent_df, on="TL_GEO_ID", how="left")
    joined_df = pd.merge(joined_df, hh_inc_df, on="TL_GEO_ID", how="left")
    joined_df = pd.merge(joined_df, renter_df, on="TL_GEO_ID", how="left")

    # Add white_pct to df
    joined_df["white_pct"] = np.where(
        joined_df["population"] > 0, joined_df["white_pop"] / joined_df["population"], 0
    )

    joined_df.to_csv(SF_ACS_JOIN, index=False)


def get_sf_geoid() -> list[str]:
    """
    Extract the list of SF census tract GeoIDs based on the list of 2020 census
    tracts from DataSF Open Data Portal

    Returns:
        List of SF census tract GeoIDs
    """
    sf_geoid = []

    csv.field_size_limit(sys.maxsize)

    with open(SF_CENSUS_PATH) as f:
        reader = csv.DictReader(f)
        for row in reader:
            geoid = row["geoid"]
            if geoid not in EXCLUDE_GEOIDS:
                sf_geoid.append(geoid)

    return sf_geoid


def create_sf_shapefiles():
    """
    Filter California census tract shapefiles and create SF census tract
    shapefiles that only include tracts in San Francisco, by matching with the
    GeoIDs obtained from the San Francisco list of census tracts.
    """
    cali_tracts = gpd.read_file(CALI_TRACTS_SHP)
    sf_geoid = get_sf_geoid()
    sf_tracts = cali_tracts[cali_tracts["GEOID"].isin(sf_geoid)]

    sf_tracts.to_file(SF_TRACTS_SHP)


def add_sf_tract_data():
    """
    Add SF ACS data to SF census tract shapefiles as attributes
    """
    sf_tracts = gpd.read_file(SF_TRACTS_SHP)
    sf_acs_data = pd.read_csv(SF_ACS_JOIN, dtype={"TL_GEO_ID": "str"})

    # Align geo_id column name in sf_acs_data with GEOID column name in
    # sf_tracts shapefile
    sf_acs_data.rename(columns={"TL_GEO_ID": "GEOID"}, inplace=True)

    updated_sf_tracts = sf_tracts.merge(sf_acs_data, on="GEOID", how="left")
    updated_sf_tracts.to_file(MERGED_SF_TRACTS_SHP)

    return sf_acs_data


def generate_zori_csv():
    """
    Loads ZORI CSV file, filters for SF zip codes and the years 2020-2024, imputes 
    to fill missing data, and outputs tidy ZORI CSV.
    
    Returns:
        zips: [list] SF zip codes
    """
    # load data
    df = pd.read_csv("raw-data/zori_by_zip.csv")

    # column(City) ==  'San Francisco'
    sf_zips = df[df["City"] == "San Francisco"].copy()

    # filter month-year(2020-2024)
    date_cols = [
        col
        for col in sf_zips.columns
        if any(yr in col for yr in ["2020", "2021", "2022", "2023", "2024"])
    ]

    filtered_df = sf_zips[["RegionName"] + date_cols]
    # change regionname to zip
    filtered_df = filtered_df.rename(columns={"RegionName": "zip"})

    # Imputes data to fill missing values
    zip_col = filtered_df["zip"]
    data = filtered_df.drop(columns=["zip"])
    data = data.interpolate(axis=1)
    data = data.fillna(data.mean())
    imputed_df = pd.concat([zip_col, data], axis=1)

    # Reformats Zori CSV into tidy format
    tidy_rows = []
    date_cols = imputed_df.drop(columns=["zip"]).columns
    for _, row in imputed_df.iterrows():
        # Avoid conversion to float by converting to string
        zip_code = str(int(row["zip"]))
        for date in date_cols:
            datetime_object = datetime.strptime(date, "%Y-%m-%d")
            formatted_date = f"{datetime_object.year}-{datetime_object.month:02}"
            tidy_rows.append({"zip": zip_code, "date": formatted_date, "rent": row[date]})

    # Writes tidy CSV
    with open("clean-data/tidy_zori.csv", "w", newline="") as f_out:
        writer = csv.DictWriter(f_out, fieldnames=["zip", "date", "rent"])
        writer.writeheader()
        writer.writerows(tidy_rows)


def generate_crosswalks_csv():
    """
    Loads crosswalks XLSX files, filters for SF zips and tracts, selects necessary 
    columns (zip, tract, res_ratio), extracts date column from filenames, and 
    outputs into single CSV.
    """
    zori_df = pd.read_csv("clean-data/tidy_zori.csv")
    acs_df = pd.read_csv("clean-data/census_acs_join.csv")
    
    zips_num = set(zori_df["zip"])
    zips = {str(zip) for zip in zips_num}
    tracts_num = set(acs_df["TL_GEO_ID"])
    short_tracts = {str(tract) for tract in tracts_num}
    tracts = {tract.zfill(11) for tract in short_tracts}

    list_of_dfs = []
    for file_path in Path("raw-data/crosswalks-xlsx").iterdir():
        if not file_path.name.startswith("~$"):
            df = pd.read_excel(file_path, engine="openpyxl")
            zip_col = None
            # Pull zip, tract, and res_ratio columns for SF zips/tracts
            for column in df.columns:
                if "zip" in column.lower():
                    zip_col = column
                    break
            df[zip_col] = df[zip_col].astype(str)            
            for column in df.columns:
                if "tract" in column.lower():
                    tract_col = column
                    break
            df[tract_col] = df[tract_col].astype(str)
            df[tract_col] = df[tract_col].str.zfill(11)
            for column in df.columns:
                if "res_ratio" in column.lower():
                    res_ratio_col = column
                    break
            # Add date column based on filename
            datetime_str = file_path.stem[-6:]
            datetime_object = datetime.strptime(datetime_str, "%m%Y")
            date = f"{datetime_object.year}-{datetime_object.month:02}"
            filtered_df = df.loc[:, [zip_col, tract_col, res_ratio_col]]
            filtered_df["date"] = date
            filtered_df.rename(
                columns={"ZIP": "zip", "TRACT": "tract", "RES_RATIO": "res_ratio"}, inplace=True
            )
            # Filter to SF zips
            sf_zips_df = filtered_df[filtered_df["zip"].isin(zips)]
            # Filter to SF tracts
            sf_df = sf_zips_df[sf_zips_df["tract"].isin(tracts)]
            list_of_dfs.append(sf_df)
    
    # Aggregate and output to CSV
    aggregated_df = pd.concat(list_of_dfs)
    aggregated_df.to_csv("clean-data/crosswalks.csv", index=None, header=True)


if __name__ == "__main__":
    process_acs_data()
    create_sf_shapefiles()
    add_sf_tract_data()
    generate_zori_csv()
    generate_crosswalks_csv()
