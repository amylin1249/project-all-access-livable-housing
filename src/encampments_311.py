##### FUNCTIONS PULLED FROM CLEAN_DATA.PY #####

import csv
import sys
import pandas as pd
import numpy as np
import geopandas as gpd
from typing import NamedTuple
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import jellyfish
import math


class Encampment(NamedTuple):
    ### unique  encampmemnt id per quarter
    id: int
    tents: int
    structures: int
    vehicles: int

    year: int
    month: int
    date_time: datetime
    lat: float
    lon: float
    neighborhood: str


class EncampmentReport(NamedTuple):
    year: int
    month: int
    address: str
    lat: float
    lon: float


def clean_parenthesis(name):
    """
    This function takes a name and removes any parenthesized portion.

    Returns:
        A string with parenthesized portion removed.
    """

    name = name.replace("(", "*(")
    name = name.replace(")", ")*")

    split_name = name.split("*")
    output_list = []

    for word in split_name:
        if word == "":
            continue
        elif word[0] != "(" and word[-1] != ")":
            output_list.append(word.strip())
    return " ".join(output_list)


STOPWORDS = [
    "st",
    "street",
    "av",
    "avenue",
    "ave",
    "av",
    "blvd",
    "boulevard",
    "rd",
    "road",
    "ln",
    "lane",
    "dr",
    "drive",
    "ct",
    "court",
    "pkwy",
    "parkway",
    "dr",
    "drive",
    "ter",
    "terrace",
    "cir",
    "circle",
    "pl",
    "place",
    "stwy",
    "a",
    "an",
    "and",
    "&",
    "are",
    "as",
    "at",
    "be",
    "by",
    "for",
    "from",
    "has",
    "he",
    "in",
    "is",
    "it",
    "its",
    "of",
    "on",
    "that",
    "the",
    "to",
    "was",
    "were",
    "will",
    "with",
    "park",
    "parks",
    "intersection",
]

PUNCTUATION = ".,?-#/()[]"


def clean_address(address):
    address = address.lower()
    address = address.replace("i poi", "")
    address = clean_parenthesis(address)
    text_data = address.split(" ")
    cleaned_list = [word.strip(PUNCTUATION) for word in text_data]
    cleaned_list = [word for word in cleaned_list if word != ""]
    cleaned_list = [word for word in cleaned_list if word not in STOPWORDS]
    return " ".join(cleaned_list)


def rate(score):
    if score >= 0.95:
        return "high"
    if score < 0.95 and score >= 0.80:
        return "medium"
    return "low"


## Clean 311 data


def clean_311():

    file_input = REPORT_PATH

    with open(file_input, newline="") as csvfile:
        """
        Given a CSV containing 311, return a list of Encampment report objects.
        """
        lat_lon_dict = {}
        output_report = []
        reader = csv.DictReader(csvfile)
        for row in reader:
            ### Clean the date ####
            datetime_str = row.get("Opened").replace(" PM", "")
            datetime_str = datetime_str.replace(" AM", "")
            datetime_object = datetime.strptime(datetime_str, "%m/%d/%Y %H:%M:%S")
            date_year = datetime_object.year
            date_month = datetime_object.month
            if row["Latitude"] == "":
                lat = 0
            else:
                lat = float(row["Latitude"])

            if row["Longitude"] == "":
                lon = 0
            else:
                lon = float(row["Longitude"])

            address = clean_address(row.get("Address"))
            tuple_out = EncampmentReport(date_year, date_month, address, None, None)
            key = tuple_out
            if key not in lat_lon_dict:
                lat_lon_dict[key] = []
                lat_lon_dict[key].append((lat, lon))

            output_report.append(tuple_out)

    return output_report, lat_lon_dict


def attach_lat_lon(output_report, lat_lon_dict):
    unique_list = set(output_report)
    output = []
    for tuple_report in list(unique_list):
        lat_lon = lat_lon_dict[tuple_report]

        lat = sum(loc[0] for loc in lat_lon if loc[0] != 0) / len(lat_lon)
        lon = sum(loc[1] for loc in lat_lon if loc[1] != 0) / len(lat_lon)

        tuple_out = EncampmentReport(
            tuple_report.year, tuple_report.month, tuple_report.address, lat, lon
        )
        output.append(tuple_out)


### Clean encampment data ###


def clean_encampment():
    file_input = ENCAMP_PATH
    wb = load_workbook(file_input)
    sheet_obj = wb.active

    for i in range(1, sheet_obj.max_column + 1):
        print(sheet_obj.cell(row=1, column=i).value)

    for i in range(1, sheet_obj.max_column + 1):
        print(sheet_obj.cell(row=2, column=i).value)

    assert sheet_obj.cell(row=2, column=3).value == "Tents"
    assert sheet_obj.cell(row=2, column=4).value == "Structures"
    assert sheet_obj.cell(row=2, column=5).value == "Passenger Vehicles"
    assert sheet_obj.cell(row=2, column=6).value == "Other Vehicles"
    assert sheet_obj.cell(row=2, column=8).value == "Neighborhood"
    assert sheet_obj.cell(row=2, column=10).value == "Latitude"
    assert sheet_obj.cell(row=2, column=11).value == "Longitude"

    output_encampment = []
    for i in range(3, sheet_obj.max_row + 1):
        sheet_obj.cell(row=3, column=1).value
        date_obj = sheet_obj.cell(row=i, column=1).value
        date_string = date_obj.strftime("%m/%d/%Y")
        tents = sheet_obj.cell(row=i, column=3).value
        structure = sheet_obj.cell(row=i, column=4).value
        vehicles = (
            sheet_obj.cell(row=i, column=5).value
            + sheet_obj.cell(row=i, column=6).value
        )
        neighborhood = sheet_obj.cell(row=i, column=8).value

        lat = float(sheet_obj.cell(row=i, column=10).value)
        lon = float(sheet_obj.cell(row=i, column=11).value)
        obj = Encampment(
            i,
            tents,
            structure,
            vehicles,
            date_obj.year,
            date_obj.month,
            date_string,
            lat,
            lon,
            neighborhood,
        )
        output_encampment.append(obj)
    return output_encampment


### Merge the two files to filter out 311 reports associatd with marked/observed encampments ####


def attached_311_reports(output_encampment, output_report):

    associated_encamp = []
    year_2021 = [encamp for encamp in output_encampment if encamp.year == 2021]
    report_2021 = [report for report in output_report if report.year == 2021]
    month_dec_2020 = [
        report for report in output_report if report.year == 2020 and report.month == 12
    ]
    month_jan_2022 = [
        report for report in output_report if report.year == 2022 and report.month == 1
    ]
    report_2021.extend(month_dec_2020)
    report_2021.extend(month_jan_2022)

    for encampment in year_2021:
        for report in report_2021:
            format_pattern = "%m/%d/%Y"
            diff = (
                datetime.strptime(encampment.date_time, format_pattern)
                - report.date_time
            )
            if abs(diff.days) <= 15:
                point1 = (encampment.lat, encampment.lon)
                point2 = (report.lat, report.lon)
                if (
                    rate(
                        jellyfish.jaro_winkler_similarity(
                            encampment.neighborhood.lower(), report.neighborhood.lower()
                        )
                    )
                    == "high"
                ):
                    if (distance.distance(point1, point2).miles) < 0.2:
                        associated_encamp.append((encampment, report))
